import pandas as pd
import numpy as np
from openpyxl import load_workbook

def append_df_to_excel(filename,
                       df,
                       sheet_name='Sheet1',
                       startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
    Returns: None
    """
    # from openpyxl import load_workbook

    # import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


df = pd.read_excel(r'D:\OneDrive - whut.edu.cn\脚本备份\MyPythonWork\testdata.xls')
data = df[df.isnull().T.any()]
k_row = data.index.values

#适用于单行模式
# j =0
# for i in range(len(k_row)):
#     if i == 0:
#         data = df[0:k_row[i]].dropna(axis=0, how='any', inplace=False)
#     else:
#         start = k_row[i-1]
#         end = k_row[i]
#         data = df[start:end].dropna(axis=0, how='any', inplace=False)
#     append_df_to_excel('save_to_execl3.xlsx',
#                        data,
#                        sheet_name='Sheet1',
#                        startcol=j,
#                        startrow=0,
#                        index=False)
#     j += 2

#适用于双行模式
j = 0
for i in range(len(k_row)):
    if i %2 == 0:
        continue
    if i == 1:
        data = df[0:k_row[i]]
    else:
        start = k_row[i - 2]
        end = k_row[i]
        data = df[start:end]
    append_df_to_excel('save_to_execl1.xlsx',
                       data,
                       sheet_name='Sheet1',
                       startcol=j,
                       startrow=0,
                       index=False)
    j += 2